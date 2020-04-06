#!/usr/bin/env python
import sys
import io
import os
import re
import pdfminer.high_level
import pdfminer.layout
import openpyxl
from openpyxl.styles import Font, Color, Alignment, Border, Side, colors, PatternFill

FILEPATH = os.path.dirname(os.path.realpath(__file__)) + '/' + './file.pdf'
SEARCH_PATTERN = r"(?:[-,.\/\\а-яА-Я0-9_\t ]+\n?)+"
LABELS_PER_COL = 5
COLUMN_WIDTH = 19.4
ROW_HEIGHT = 60
FONT_SIZE = 8
FONT_NAME = "Arial" # Arial, Times New Roman
SIDES_TEXT = "Гарнитура"

def export_labels(pdf: str) -> list:
    print("Attempting to parse pdf file:", pdf)
    with open(pdf, 'rb') as f:
        output_string = io.StringIO()
        laparams = pdfminer.layout.LAParams() # default layout parameters
        try:
            pdfminer.high_level.extract_text_to_fp(f, output_string, laparams=laparams,
                page_numbers=list(range(1, 30)))

            print("Parsed pdf file successfully.")
        except: # usually pdfminer.pdfparser.PDFSyntaxError
            print("Could not parse pdf file:", pdf)
        
        print("Extracting labels...")
        string = output_string.getvalue().strip()
        pattern = re.compile(SEARCH_PATTERN)
        matches = re.finditer(pattern, string)
        labels = []
        count = 0
        for match in matches:
            print(count, match)
            #label = " ".join(match.group(0).strip().split())
            label = match.group(0).strip()
            labels.append(label)
            count += 1
        return labels

def clear_whitespace(items: list) -> list:
    result = []
    for item in items:
        result.append(" ".join(item.split()))
    return result

def remove_empty_strings(words: list) -> list:
    result = []
    for word in words:
        if word:
            result.append(word)
    return result

#
# Separate the sides into a new list
#
def separate_sides(labels: list) -> list:
    sides = []
    for label in labels.copy():
        if SIDES_TEXT.lower() in label.lower():
            #print("Found side:", label)
            sides.append(label)
            labels.remove(label)
    return sides

#
# Split a label into a list words
#
def split_into_words(item: str) -> list:
    delimiters = " ", "\n", "\t"
    regexPattern = '|'.join(map(re.escape, delimiters))
    words = remove_empty_strings(re.split(regexPattern, item))
    return words

#
# Extract the last 3 words from a list of words
#
def get_client_names(words: list) -> list:
    length = len(words)
    if length < 3:
        raise ValueError("List of words contains less than 3 words. Can't extract client names.")

    i = 3
    names = []
    while i > 0:
        names.append(words[length-i])
        i -= 1
    
    return names

#
# Append the sides to a more appropriate position in the original list of labels
#
def reinsert_sides(sides: list, labels: list):
    for side in sides:
        # extract the name of the client
        side_words = split_into_words(side)
        side_names = get_client_names(side_words)
        #print("side_words:", side_words)
        #print("side_names:", side_names)

        # Find an appropriate position
        for i, label in enumerate(labels.copy()):
            label_words = split_into_words(label)
            #print("label_words:", label_words)
            # do the side names exist in this label?
            has_names =  all(elem in label_words  for elem in side_names)
            if has_names is True:
                print("Found a matching position at", i + 1, "for:", " ".join(side.split()))
                labels.insert(i + 1, side)
                break
        else:
            print("Could not find an appropriate position for:", side)

    return

def create_spreadsheet(labels: list, filename: str):
    print("Creating spreadsheet...")
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # a bit of styling
    center_aligned_text = Alignment(horizontal="center",
                                     vertical='center',
                                     wrap_text=True)

    for col in ['A', 'B', 'C', 'D', 'E']:
        sheet.column_dimensions[col].width = COLUMN_WIDTH
    
    grayFill = PatternFill(fgColor=Color(rgb="dddddddd"),
                   fill_type='solid')

    # populate the sheet
    column, row = 0, 0
    for i, label in enumerate(labels):
        if (column := i % LABELS_PER_COL) == 0:
            row = row + 1
            sheet.row_dimensions[row].height = ROW_HEIGHT

        sheet.cell(row, column + 1, label).alignment = center_aligned_text
        sheet.cell(row, column + 1).font = Font(size = FONT_SIZE, name = FONT_NAME)

        if SIDES_TEXT.lower() in label.lower():
            sheet.cell(row, column + 1).fill = grayFill
    
    # fix margins
    sheet.page_margins.left = 0.39
    sheet.page_margins.right = 0.39
    sheet.page_margins.top = 0.39
    sheet.page_margins.bottom = 0.39
    sheet.page_margins.header = 0
    sheet.page_margins.footer = 0
    sheet.print_options.horizontalCentered = True
    
    try:
        filename = filename.replace(".pdf", ".xlsx")
        workbook.save(filename = filename)
        print("Saved spreadsheet:", filename)
    except PermissionError:
        print("Could not open file for writing:", filename)
        print("Please make sure the file is not being read by another program.")
    except:
        print("Failed to save spreadsheet:", filename)

def main(args: list):
    if len(args) < 2:
        print('Usage: ./extract-labels <pdffile1> <pdffile2> ...')
        input("Press ENTER to exit.")
        return
    
    for i, a in enumerate(args):
        if i == 0:
            continue
        labels = export_labels(a)
        print("Number of labels before:", len(labels))

        # separate sides
        sides = separate_sides(labels)
        print("Гарнитури (", len(sides), ")")
        for i, side in enumerate(clear_whitespace(sides)):
            print(i, side)

        reinsert_sides(sides, labels)

        print("Number of labels after:", len(labels))
        for i, label in enumerate(clear_whitespace(labels)):
            print(i, label)

        create_spreadsheet(labels, a)
    
    input("Job finished.")

if __name__ == '__main__':
    main(sys.argv)
